#!/usr/bin/python3

import openpyxl
import os
import shutil

def get_brand_model_dict ():
    
    refer_excel = openpyxl.load_workbook('language.xlsx')
    refer_sheet = refer_excel['Language']
    brand_model_dict = {}
    
    folder = os.getcwd()
    
    for item in range(2, refer_sheet.max_column+1):
        name = (refer_sheet.cell(row=1, column=item)).value
        filePath = folder + "/" + name + ".txt"
        if os.path.exists(filePath):
            os.remove(filePath)
        
        suffix = 0
        for row in range(2, refer_sheet.max_row+1):
            brand = (refer_sheet.cell(row=row, column=3)).value
            model = (refer_sheet.cell(row=row, column=item)).value
        
            if brand is None:
                print(" ")
            elif brand:
                if model is None:
                    print(" ")
                elif model:
                    file = open(filePath, 'a')
                    bt1 = "\""+brand+"\""
                    suffix = 0
                    for item_1 in open(filePath):
                        tmp = item_1.find(bt1)
                        #print("item_1.find",tmp,item_1,bt1)
                        if tmp != -1:
                            suffix += 1
                            print(bt1,suffix)
                            brand = brand +"_"+str(suffix)
                    print("brand：", brand)
                    file.write("\""+brand+"\""+" = "+"\""+model+"\""";\n")
                    file.close()
                
            
    

# file = open(filePath, 'a')
# if brand in open(filePath).read():
#     bt1 = "\""+brand+"\""
    
#     while bt1 in open(filePath).readline():
#         suffix += 1
#         print(bt1,suffix)

#     brand = brand +"_"+str(suffix)
#     print("brand：", brand)
# file.write("\""+brand+"\""+" = "+"\""+model+"\""";\n")
# file.close()

def copyToTarget (source_file,destination_file):

    with open(source_file, "r", encoding="utf-8") as f:
        content = f.read()

    if os.path.exists(destination_file):
            os.remove(destination_file)  

    with open(destination_file, "w", encoding="utf-8") as f:
        f.write(content)
    
    shutil.copy2(source_file, destination_file)

    

     

# ~~~~~~~~~~~~~~~~main~~~~~~~~~~~~~~~~~~
# 读取对应关系表格：

def test ():
    get_brand_model_dict()
    folder = os.getcwd()
    list = os.listdir(folder)
    parent_path = os.path.dirname(folder)+"/SDKTestHelper/SDKTestHelper/Sources/"

    for item in list:
        if item.endswith(".txt"):
            file_path = os.path.join(folder, item)
            targetPath = ""
            if item == "英语.txt":
                targetPath = parent_path+"en.lproj/"
            if item == "中文.txt":
                targetPath = parent_path+"zh-Hans.lproj/"

            copyToTarget(
                    file_path, targetPath+"Localizable.strings")
            
        
                



test()



    
